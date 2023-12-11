# Construa um programa em python e apresente no terminal, as seguintes respostas:
# a) A listagem dos vendedores seguido da soma de suas comissões
# b) Apresente o vendedor que mais ganhou comissões
# c) Apresente o vendedor que mais ganhou comissões no mês de Junho
# d) Liste todas as vendas da vendedora Alice
# e) Liste todas as vendas separadas para cada vendedor

import openpyxl

caminho = openpyxl.load_workbook('Exercício de Algoritmos Controle de Vendas.xlsx', data_only=True)
planilha = caminho['Página1']

comissaoAlice = 0
comissaoArthur = 0
comissaoMatheus = 0
comissaoAliceJunho = 0
comissaoArthurJunho = 0
comissaoMatheusJunho = 0

for linha in range(5, 23):
    vendedor = planilha.cell(row=linha, column=6).value
    comissao = planilha.cell(row=linha, column=5).value * planilha.cell(row=linha, column=7).value
    dataVenda = planilha.cell(row=linha, column=1).value

    if vendedor == 'Alice':
        comissaoAlice += comissao
        if dataVenda.month == 6:
            comissaoAliceJunho += comissao

    if vendedor == 'Arthur':
        comissaoArthur += comissao
        if dataVenda.month == 6:
            comissaoArthurJunho += comissao

    if vendedor == 'Matheus':
        comissaoMatheus += comissao
        if dataVenda.month == 6:
            comissaoMatheusJunho += comissao

print(f'Comissão Alice: R$ {comissaoAlice}')
print(f'Comissão Matheus: R$ {comissaoMatheus}')
print(f'Comissão Arthur: R$ {comissaoArthur}')

if comissaoArthur > comissaoAlice:
    maiorComissao = comissaoArthur
    nome = 'Arthur'

else:
    maiorComissao = comissaoAlice
    nome = 'Alice'

if maiorComissao < comissaoMatheus:
    maiorComissao = comissaoMatheus
    nome = 'Matheus'

print(f'Vendedor com maior comissão: {nome}, valor: R$ {maiorComissao}')

if comissaoArthurJunho > comissaoAliceJunho:
    comissaoJunho = comissaoArthurJunho
    nome = 'Arthur'

else:
    comissaoJunho = comissaoAliceJunho
    nome = 'Alice'

if comissaoJunho < comissaoMatheusJunho:
    comissaoJunho = comissaoMatheusJunho
    nome = 'Matheus'

print(f'Vendedor com maior comissão em junho: {nome}, valor: R$ {comissaoJunho}')

# d)
for linha in range(5, 31):
    vendedor = planilha.cell(row=linha, column=6).value
    if vendedor == 'Alice':
        for celula in planilha[f"A{linha}:H{linha}"][0]:
            print(celula.value)

# linha a linha
for linha in range(5, 31):
    vendedor = planilha.cell(row=linha, column=6).value
    if vendedor == 'Alice':
        valores_linha = [planilha.cell(row=linha, column=coluna).value for coluna in range(1, 9)]
        print(valores_linha)


print('---'*15)

# e) organizando os dados por vendedor
# Dicionário para armazenar os dados por vendedor
dados_por_vendedor = {}

# Percorre a planilha para coletar os dados por vendedor
for linha in range(5, 23):
    vendedor = planilha.cell(row=linha, column=6).value
    if vendedor not in dados_por_vendedor:
        dados_por_vendedor[vendedor] = []

    valores_linha = [planilha.cell(row=linha, column=coluna).value for coluna in range(1, 9)]
    dados_por_vendedor[vendedor].append(valores_linha)

# Imprime os dados organizados por vendedor
for vendedor, dados in dados_por_vendedor.items():
    print(f"Vendedor: {vendedor}")
    for linha in dados:
        print(linha)
    print("\n")
