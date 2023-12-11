import openpyxl
caminho = openpyxl.load_workbook('Planilha de Frutas.xlsx')

# Para colocar a planilha na aba escolhida
planilha = caminho['Frutas']

# Ler  célula a célula de cada linha

#     Para cada linha na aba Frutas - da linha 2 à linha 5
for rows in planilha.iter_rows(min_row=2, max_row=5):
    #   Para cada célula na linha:
    for cell in rows:
        # troca de valor de uma célula
        if cell.value == 'Banana':
            cell.value = 'Tamarindo'
        #  Imprimir o valor da célula.
        print(cell.value)

print('----'*3)

# Ler os dados da linha da planilha (linha inteira)

for rows in planilha.iter_rows(max_row=5):
    print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}')


print('----'*3)
# Outra maneira de imprimir os dados da planilha
for rows in planilha.iter_rows(max_row=5):
    print(rows[0].value, rows[1].value, rows[2].value)

# Salvando a planilha
caminho.save('Planilha de Frutas.xlsx')

# fruta= []
# for linha in range(2, 6):
#     valorCelula = planilha.cell(row=linha, column=1).value
#     fruta.append(valorCelula)
#     print(planilha.cell(row=linha, column=1).value)
#     print(planilha[f'A{linha}'].value)

# print(fruta)

# for coluna in range(1, 4):
#     print(planilha.cell(row=2, column=coluna).value)


# for columns in planilha.iter_cols(max_col=4):
#     for cell in columns:
#         if cell.value == 'Tamarindo':
#             print('Felipe')

# for i in range(1, 5):
#     for j in range(1, 7):
#         celula = planilha.cell(row=i, column=j)  #Aqui é o pulo do gato usando range. Acessar as celulas.
#         if celula.value == 'Tamarindo':
#             print('Felipe')
