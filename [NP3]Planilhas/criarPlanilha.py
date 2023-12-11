import openpyxl

# Criar uma planilha no excel
planilha = openpyxl.Workbook()

# Como visualizar páginas existentes
print(planilha.sheetnames)

# Como criar uma página
planilha.create_sheet('Frutas')

# Como selecionar uma página
planilhaDeFrutas = planilha['Frutas']

# Adicionar dados à planilha. Cada item da lista [], será uma célula da planilha na mesma linha
planilhaDeFrutas.append(['Fruta', 'Quantidade', 'Preço'])
planilhaDeFrutas.append(['Banana', '5', 'R$ 3,90'])
planilhaDeFrutas.append(['Maçã', '6', 'R$ 5,20'])
planilhaDeFrutas.append(['Mamão', '1', 'R$ 3,40'])
planilhaDeFrutas.append(['Limão', '12', 'R$ 6,00'])

# Salvando a planilha
planilha.save('Planilha de Frutas.xlsx')
