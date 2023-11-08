
from openpyxl import load_workbook

#https://www.datacamp.com/tutorial/python-excel-tutorial
#https://openpyxl.readthedocs.io/en/stable/tutorial.html
#https://docs.google.com/spreadsheets/d/1r4lHlaj7Mh-TD7KUrE9SQx-Kp-Yz374L/edit?usp=sharing&ouid=111182560362271475427&rtpof=true&sd=true

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active
conteudo = planilha['B5']
celula = []
celula.append(conteudo.value)

planilha.cell(row=7, column=4, value='')
arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
alunos = []
for i in range(6, 16):
    conteudo = planilha[f'B{i}']
    alunos.append(conteudo.value)

resultadoAlunosSim = {}
resultadoAlunosNao = {}

for a in alunos:
    somasim = 0
    somanao = 0
    for i in range(6, 45):
        celulaAluno = planilha[f'B{i}']
        celulaResposta = planilha[f'C{i}']
        if celulaResposta.value != None:
            if celulaAluno.value == a and celulaResposta.value == 'SIM':
                somasim += 1
                resultadoAlunosSim[a] = somasim
            elif celulaAluno.value == a and (celulaResposta.value == 'NÃO' or celulaResposta.value == 'NAO'):
                somanao += 1
                resultadoAlunosNao[a] = somanao
print("SIM", resultadoAlunosSim)
print("NÃO", resultadoAlunosNao)


indice = 0
for z in range(49, 49 + len(alunos)):
        planilha.cell(row=z, column=5, value=alunos[indice])
        indice += 1

arquivo_excel.save("Aprendendoalerplanilhas.xlsx")

indice = 0
for z in range(49, 49 + len(alunos)):
    if alunos[indice] == planilha[f'E{z}'].value:
        planilha.cell(row=z, column=6, value=resultadoAlunosSim.get(alunos[indice]))
        planilha.cell(row=z, column=7, value=resultadoAlunosNao.get(alunos[indice]))
        indice += 1


    # Esse código Python trabalha com planilhas no formato Excel usando a biblioteca openpyxl.
    # Vou explicar cada parte:
    #
    # caminho = 'Aprendendoalerplanilhas.xlsx': Define o caminho do arquivo Excel que será lido e modificado.
    #
    # arquivo_excel = load_workbook(caminho, data_only=True): Carrega o arquivo Excel especificado em caminho.
    # O argumento data_only=True indica que os valores das células devem ser carregados como valores (números ou strings),
    # não como fórmulas.
    #
    # planilha = arquivo_excel.active: Acessa a planilha ativa do arquivo carregado.
    #
    # conteudo = planilha['B5']: Lê o conteúdo da célula B5.
    #
    # celula = []: Inicializa uma lista vazia chamada celula.
    #
    # celula.append(conteudo.value): Adiciona o valor da célula B5 à lista celula.
    #
    # planilha.cell(row=7, column=4, value=''): Limpa o conteúdo da célula na linha 7, coluna 4.
    #
    # arquivo_excel.save("Aprendendoalerplanilhas.xlsx"): Salva as modificações no arquivo Excel.
    #
    # alunos = []: Inicializa uma lista vazia chamada alunos.
    #
    # O loop for i in range(6, 16) percorre as linhas de 6 a 15 da coluna B e adiciona os valores à lista alunos.
    #
    # resultadoAlunosSim = {} e resultadoAlunosNao = {}: Inicializam dois dicionários vazios para armazenar os resultados de "SIM" e "NÃO".
    #
    # Os loops aninhados percorrem as linhas de 6 a 44.
    # Para cada linha, verifica se a célula na coluna B (referente ao aluno) e a célula na coluna C (referente à resposta) possuem valores.
    # Se sim, verifica se a resposta é "SIM" ou "NÃO" e incrementa o contador correspondente no dicionário de resultados.
    #
    # print("SIM", resultadoAlunosSim) e print("NÃO", resultadoAlunosNao): Imprime os resultados dos alunos que responderam "SIM" e "NÃO".
    #
    # Os próximos loops adicionam os nomes dos alunos e os resultados "SIM" e "NÃO" nas colunas E, F e G, respectivamente, a partir da linha 49.
    #
    # Por fim, o arquivo é salvo novamente com as modificações.