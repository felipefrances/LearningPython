# Construa um programa em python e apresente no terminal, as seguintes respostas:
# a) A listagem dos vendedores seguido da soma de suas comissões
# b) Apresente o vendedor que mais ganhou comissões
# c) Apresente o vendedor que mais ganhou comissões no mês de Junho

# https://docs.google.com/spreadsheets/d/1U1XDkHXGwVWOR78kJbSO82je6w4ooM94VSMuJiGymVc/edit#gid=0



from openpyxl import load_workbook

caminho = 'Exercício de Algoritmos Controle de Vendas.xlsx'

arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active

aliceComissao = 0
arthurComissao = 0
matheusComissao = 0

aliceComissaoJunho = 0
arthurComissaoJunho = 0
matheusComissaoJunho = 0

maiorComissao = 0

maiorComissaoJunho = 0

linha_atual = 5

while linha_atual <= 22:
    funcionario = planilha.cell(row=linha_atual, column=6).value #captura os nomes dos funcionarios
    comissao = planilha.cell(row=linha_atual, column=8).value #captura as comissões dos funcionarios

    if funcionario == "Alice":
        aliceComissao += comissao
    elif funcionario == "Arthur":
        arthurComissao += comissao
    elif funcionario == "Matheus":
        matheusComissao += comissao

    mes = planilha.cell(row=linha_atual, column=1).value #percorre a data da planilha

    mesJunho = mes.month #atribui o mes a variavel mesJunho com usando a função month captura apenas o mês da data
    if mesJunho == 6:
        if funcionario == "Alice":
            aliceComissaoJunho += comissao
        elif funcionario == "Arthur":
            arthurComissaoJunho += comissao
        elif funcionario == "Matheus":
            matheusComissaoJunho += comissao

    linha_atual += 1


print(f'A soma da comissão de Alice foi de R${aliceComissao}')
print(f'A soma da comissão de Arthur foi de R${arthurComissao}')
print(f'A soma da comissão de Matheus foi de R${matheusComissao}')

if aliceComissao > arthurComissao and aliceComissao > matheusComissao:
    maiorComissao = aliceComissao
    print(f'O vendedor com maior comissão foi a Alice com R${maiorComissao}')
elif arthurComissao > aliceComissao and arthurComissao > matheusComissao:
    maiorComissao = arthurComissao
    print(f'O vendedor com maior comissão foi o Arthur com R${maiorComissao}')
elif matheusComissao > aliceComissao and matheusComissao > arthurComissao:
    maiorComissao = matheusComissao
    print(f'O vendedor com maior comissão foi o Matheus com R${maiorComissao}')


if aliceComissaoJunho > arthurComissaoJunho and aliceComissaoJunho > matheusComissaoJunho:
    maiorComissaoJunho = aliceComissaoJunho
    print(f'O vendedor que mais ganhou comissões no mês de Junho foi a Alice com R${maiorComissaoJunho}')
elif arthurComissaoJunho > aliceComissaoJunho and arthurComissaoJunho > matheusComissaoJunho:
    maiorComissaoJunho = arthurComissaoJunho
    print(f'O vendedor que mais ganhou comissões no mês de Junho foi o Arthur com R${maiorComissaoJunho}')
elif matheusComissaoJunho > aliceComissaoJunho and matheusComissaoJunho > arthurComissaoJunho:
    maiorComissaoJunho = matheusComissaoJunho
    print(f'O vendedor que mais ganhou comissões no mês de Junho foi o Matheus com R${maiorComissaoJunho}')